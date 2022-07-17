import pandas as pd
from openpyxl import load_workbook
from rest_framework.exceptions import NotFound

from arabic_writers.serializers import DataSerializer


class SheetFile(object):

    @staticmethod
    def list_xlsx_file_data(xlsx_file_path):
        data_frame = pd.read_excel(xlsx_file_path, engine='openpyxl',
                                   dtype=object, header=None)
        return DataSerializer(data_frame.values.tolist()[1:], many=True).data

    @staticmethod
    def add_data_into_xlsx_file(xlsx_file_path, new_data):
        workbook = load_workbook(xlsx_file_path)
        worksheet = workbook.worksheets[0]
        worksheet.append(new_data)
        workbook.save(xlsx_file_path)

    @staticmethod
    def retrieve_xlsx_file_data(xlsx_file_path, row_data):
        row_serializer = DataSerializer()

        data_frame = pd.read_excel(xlsx_file_path, engine='openpyxl',
                                   dtype=object, header=None)
        for row in data_frame.values.tolist()[1:]:
            if str(row[0]).strip('\n') == str(row_data):
                row_serializer = DataSerializer(row)
                break
        return row_serializer.data

    @staticmethod
    def update_xlsx_file_data(xlsx_file_path, row_data, new_data):
        workbook = load_workbook(xlsx_file_path)
        sheet1 = workbook.get_sheet_by_name('Sheet1')

        for col in range(1, len(new_data) + 1):
            cell = sheet1.cell(row=row_data, column=col)
            if not cell.value:
                raise NotFound

            cell.value = new_data[col-1]

        workbook.save(xlsx_file_path)

    @staticmethod
    def delete_xlsx_file_data(xlsx_file_path, row_data):
        workbook = load_workbook(xlsx_file_path)
        worksheet = workbook.worksheets[0]
        worksheet.delete_rows(row_data)
        workbook.save(xlsx_file_path)
